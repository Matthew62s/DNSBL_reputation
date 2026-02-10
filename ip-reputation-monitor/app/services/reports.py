"""Report generation service (CSV, XLSX, PDF)."""

import csv
import io
import json
import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak

from app.core.config import settings
from app.core.database import get_db_context
from app.models.database import CheckResult, MonitorRun, Target, Zone


class ReportService:
    """Service for generating reports."""

    def __init__(self):
        self.reports_dir = settings.REPORTS_DIR
        os.makedirs(self.reports_dir, exist_ok=True)

    def generate_report(
        self,
        report_type: str,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        target_ids: Optional[List[int]] = None,
        zone_ids: Optional[List[int]] = None,
        status_filter: Optional[str] = None,
    ) -> tuple[str, int]:
        """
        Generate a report and return file path and size.

        Returns:
            Tuple of (file_path, file_size_bytes)
        """
        # Gather data
        data = self._gather_report_data(date_from, date_to, target_ids, zone_ids, status_filter)

        # Generate based on type
        if report_type == "csv":
            return self._generate_csv(data)
        elif report_type == "xlsx":
            return self._generate_xlsx(data)
        elif report_type == "pdf":
            return self._generate_pdf(data)
        else:
            raise ValueError(f"Unsupported report type: {report_type}")

    def _gather_report_data(
        self,
        date_from: Optional[datetime] = None,
        date_to: Optional[datetime] = None,
        target_ids: Optional[List[int]] = None,
        zone_ids: Optional[List[int]] = None,
        status_filter: Optional[str] = None,
    ) -> dict:
        """Gather data for the report."""
        with get_db_context() as db:
            # Report scope: always use latest monitor run results.
            latest_run = (
                db.query(MonitorRun)
                .filter(MonitorRun.status == "completed")
                .order_by(MonitorRun.started_at.desc())
                .first()
            )

            if not latest_run:
                latest_run = db.query(MonitorRun).order_by(MonitorRun.started_at.desc()).first()

            # Build query
            query = db.query(CheckResult, Target, Zone).join(Target, CheckResult.target_id == Target.id).join(Zone, CheckResult.zone_id == Zone.id)

            if latest_run:
                query = query.filter(CheckResult.run_id == latest_run.id)
            else:
                query = query.filter(CheckResult.id == -1)

            # Apply target filter
            if target_ids:
                query = query.filter(CheckResult.target_id.in_(target_ids))

            # Apply zone filter
            if zone_ids:
                query = query.filter(CheckResult.zone_id.in_(zone_ids))

            # Apply status filter
            if status_filter and status_filter != "all":
                query = query.filter(CheckResult.status == status_filter)

            # Execute query
            # NOTE: keep only the most recent result for each (target, zone) pair
            # so reports do not duplicate entries when multiple checks were stored
            # for the same target/zone in a single monitor run.
            raw_results = query.order_by(CheckResult.last_checked.desc()).all()
            seen_pairs = set()
            results = []
            for result in raw_results:
                r, t, z = result
                pair_key = (t.id, z.id)
                if pair_key in seen_pairs:
                    continue
                seen_pairs.add(pair_key)
                results.append(result)

            # Process data
            summary = {
                "total_results": len(results),
                "listed_count": sum(1 for r, t, z in results if r.status == "listed"),
                "blocked_count": sum(1 for r, t, z in results if r.status == "blocked"),
                "error_count": sum(1 for r, t, z in results if r.status == "error"),
                "blocked_ip_addresses": sorted({
                    t.target for r, t, z in results if r.status == "blocked" and t.type == "ip"
                }),
                "monitor_run_id": latest_run.id if latest_run else None,
                "monitor_run_started_at": latest_run.started_at.isoformat() if latest_run else None,
                "monitor_run_finished_at": latest_run.finished_at.isoformat() if latest_run and latest_run.finished_at else None,
            }

            # Per-target breakdown
            target_breakdown = {}
            for r, t, z in results:
                if t.id not in target_breakdown:
                    target_breakdown[t.id] = {
                        "target": t.target,
                        "type": t.type,
                        "label": t.label,
                        "listed": 0,
                        "blocked": 0,
                        "errors": 0,
                    }
                if r.status == "listed":
                    target_breakdown[t.id]["listed"] += 1
                elif r.status == "blocked":
                    target_breakdown[t.id]["blocked"] += 1
                elif r.status == "error":
                    target_breakdown[t.id]["errors"] += 1

            # Per-zone breakdown
            zone_breakdown = {}
            for r, t, z in results:
                if z.id not in zone_breakdown:
                    zone_breakdown[z.id] = {
                        "zone": z.zone,
                        "listed": 0,
                        "blocked": 0,
                        "errors": 0,
                    }
                if r.status == "listed":
                    zone_breakdown[z.id]["listed"] += 1
                elif r.status == "blocked":
                    zone_breakdown[z.id]["blocked"] += 1
                elif r.status == "error":
                    zone_breakdown[z.id]["errors"] += 1

            # Detailed results
            detailed_results = []
            for r, t, z in results:
                detailed_results.append({
                    "target": t.target,
                    "target_type": t.type,
                    "zone": z.zone,
                    "status": r.status,
                    "a_records": json.loads(r.a_records) if r.a_records else [],
                    "error_reason": r.error_reason,
                    "last_checked": r.last_checked.isoformat(),
                    "last_seen": r.last_seen.isoformat(),
                })

            return {
                "summary": summary,
                "target_breakdown": list(target_breakdown.values()),
                "zone_breakdown": list(zone_breakdown.values()),
                "detailed_results": detailed_results,
            }

    def _generate_csv(self, data: dict) -> tuple[str, int]:
        """Generate CSV report."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.csv"
        filepath = os.path.join(self.reports_dir, filename)

        with open(filepath, "w", newline="") as f:
            writer = csv.writer(f)

            # Write summary
            writer.writerow(["Summary"])
            writer.writerow(["Total Results", data["summary"]["total_results"]])
            writer.writerow(["Listed Count", data["summary"]["listed_count"]])
            writer.writerow(["Blocked Count", data["summary"]["blocked_count"]])
            writer.writerow([
                "Blocked IP Addresses",
                ", ".join(data["summary"]["blocked_ip_addresses"])
                if data["summary"]["blocked_ip_addresses"]
                else "None",
            ])
            writer.writerow(["Error Count", data["summary"]["error_count"]])
            writer.writerow(["Monitor Run ID", data["summary"]["monitor_run_id"] or "N/A"])
            writer.writerow(["Run Started", data["summary"]["monitor_run_started_at"] or "N/A"])
            writer.writerow(["Run Finished", data["summary"]["monitor_run_finished_at"] or "N/A"])
            writer.writerow([])

            # Write detailed results
            writer.writerow(["Detailed Results"])
            writer.writerow(["Target", "Type", "Zone", "Status", "A Records", "Error", "Last Checked", "Last Seen"])

            for result in data["detailed_results"][:settings.REPORT_MAX_ROWS]:
                writer.writerow([
                    result["target"],
                    result["target_type"],
                    result["zone"],
                    result["status"],
                    json.dumps(result["a_records"]),
                    result["error_reason"] or "",
                    result["last_checked"],
                    result["last_seen"],
                ])

        file_size = os.path.getsize(filepath)
        return filepath, file_size

    def _generate_xlsx(self, data: dict) -> tuple[str, int]:
        """Generate XLSX report."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)

        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Write summary
        ws_summary.append(["Summary"])
        ws_summary.append(["Total Results", data["summary"]["total_results"]])
        ws_summary.append(["Listed Count", data["summary"]["listed_count"]])
        ws_summary.append(["Blocked Count", data["summary"]["blocked_count"]])
        ws_summary.append([
            "Blocked IP Addresses",
            ", ".join(data["summary"]["blocked_ip_addresses"])
            if data["summary"]["blocked_ip_addresses"]
            else "None",
        ])
        ws_summary.append(["Error Count", data["summary"]["error_count"]])
        ws_summary.append(["Monitor Run ID", data["summary"]["monitor_run_id"] or "N/A"])
        ws_summary.append(["Run Started", data["summary"]["monitor_run_started_at"] or "N/A"])
        ws_summary.append(["Run Finished", data["summary"]["monitor_run_finished_at"] or "N/A"])

        # Format summary header
        ws_summary["A1"].font = Font(bold=True, size=14)

        # Create target breakdown sheet
        ws_targets = wb.create_sheet("Target Breakdown")
        ws_targets.append(["Target", "Type", "Label", "Listed", "Blocked", "Errors"])

        for item in data["target_breakdown"]:
            ws_targets.append([
                item["target"],
                item["type"],
                item["label"] or "",
                item["listed"],
                item["blocked"],
                item["errors"],
            ])

        # Format header
        for cell in ws_targets[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Create zone breakdown sheet
        ws_zones = wb.create_sheet("Zone Breakdown")
        ws_zones.append(["Zone", "Listed", "Blocked", "Errors"])

        for item in data["zone_breakdown"]:
            ws_zones.append([
                item["zone"],
                item["listed"],
                item["blocked"],
                item["errors"],
            ])

        # Format header
        for cell in ws_zones[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Create detailed results sheet
        ws_details = wb.create_sheet("Detailed Results")
        ws_details.append(["Target", "Type", "Zone", "Status", "A Records", "Error", "Last Checked", "Last Seen"])

        for result in data["detailed_results"][:settings.REPORT_MAX_ROWS]:
            ws_details.append([
                result["target"],
                result["target_type"],
                result["zone"],
                result["status"],
                json.dumps(result["a_records"]),
                result["error_reason"] or "",
                result["last_checked"],
                result["last_seen"],
            ])

        # Format header
        for cell in ws_details[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

        # Adjust column widths
        for sheet in wb:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width

        wb.save(filepath)

        file_size = os.path.getsize(filepath)
        return filepath, file_size

    def _generate_pdf(self, data: dict) -> tuple[str, int]:
        """Generate PDF report."""
        timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        filename = f"report_{timestamp}.pdf"
        filepath = os.path.join(self.reports_dir, filename)

        doc = SimpleDocTemplate(filepath, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = styles["Heading1"]
        title = Paragraph("IP Reputation Monitor Report", title_style)
        story.append(title)
        story.append(Spacer(1, 12))

        # Summary
        summary_style = styles["Normal"]
        story.append(Paragraph("<b>Summary</b>", styles["Heading2"]))
        story.append(Paragraph(f"Total Results: {data['summary']['total_results']}", summary_style))
        story.append(Paragraph(f"Listed Count: {data['summary']['listed_count']}", summary_style))
        story.append(Paragraph(f"Blocked Count: {data['summary']['blocked_count']}", summary_style))
        blocked_ip_addresses = data["summary"]["blocked_ip_addresses"]
        blocked_ip_text = ", ".join(blocked_ip_addresses) if blocked_ip_addresses else "None"
        story.append(Paragraph(f"Blocked IP Addresses: {blocked_ip_text}", summary_style))
        story.append(Paragraph(f"Error Count: {data['summary']['error_count']}", summary_style))
        story.append(Paragraph(f"Date From: {data['summary']['date_from'] or 'N/A'}", summary_style))
        story.append(Paragraph(f"Date To: {data['summary']['date_to'] or 'N/A'}", summary_style))
        story.append(Spacer(1, 12))

        # Zone breakdown table
        story.append(Paragraph("<b>Zone Breakdown</b>", styles["Heading2"]))
        zone_data = [["Zone", "Listed", "Blocked", "Errors"]]
        for item in data["zone_breakdown"]:
            zone_data.append([
                item["zone"],
                str(item["listed"]),
                str(item["blocked"]),
                str(item["errors"]),
            ])
        zone_table = Table(zone_data)
        zone_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(zone_table)
        story.append(Spacer(1, 12))

        # Build PDF
        doc.build(story)

        file_size = os.path.getsize(filepath)
        return filepath, file_size

    def cleanup_old_reports(self, retention_days: int = None) -> int:
        """Remove reports older than retention period."""
        retention_days = retention_days or settings.REPORT_RETENTION_DAYS
        cutoff = datetime.utcnow().timestamp() - (retention_days * 24 * 60 * 60)

        removed_count = 0
        for filename in os.listdir(self.reports_dir):
            filepath = os.path.join(self.reports_dir, filename)
            if os.path.isfile(filepath):
                file_mtime = os.path.getmtime(filepath)
                if file_mtime < cutoff:
                    os.remove(filepath)
                    removed_count += 1

        return removed_count


# Global report service instance
_report_service: Optional[ReportService] = None


def get_report_service() -> ReportService:
    """Get or create global report service instance."""
    global _report_service
    if _report_service is None:
        _report_service = ReportService()
    return _report_service
